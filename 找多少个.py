import pandas as pd
from openpyxl import load_workbook


# 读取原始Excel文件
file_path = "数据整合\鸟撞有建筑合集_20250319_110911.xlsx"
df = pd.read_excel(file_path)



required_columns = [
    '4.鸟撞发生处周边环境', '5.鸟撞发生状况', '7.撞击面玻璃占比', '8.撞击面方向', 
    '10.此建筑总共有几层楼？', '11.此建筑总体上玻璃的覆盖比例为（%）：', 
    '12.此建筑总体上有防鸟撞措施覆盖的玻璃比例为（%）：', '15.此建筑周围五米内占比最多的环境类型是', 
    '12.鸟种鉴定'
]

result_columns = [
    '4.鸟撞发生处周边环境', '5.鸟撞发生状况', '7.撞击面玻璃占比', '8.撞击面方向', 
    '10.此建筑总共有几层楼', '11.此建筑总体上玻璃的覆盖比例为', 
    '12.防鸟撞措施覆盖的玻璃比例为', '15.周围五米内占比最多的环境类型', 
    '12.鸟种鉴定'
]

a=len(required_columns)
print(a)

for i in range(a):
    # 替换 '列名' 为你要处理的具体列
    column_name = required_columns[i]
    output_name = result_columns[i]

    # 统计唯一值及其计数
    count_data = df[column_name].value_counts().reset_index() 
    count_data.columns = [column_name, 'Count']   #将 DataFrame 的列名重命名

    # 检查已有工作表名
    workbook = load_workbook(file_path)
    sheet_name = output_name 
    if sheet_name in workbook.sheetnames:
        i = 1
        while f"{output_name}_{i}" in workbook.sheetnames:
            i += 1
        sheet_name = f"{output_name}_{i}"
    
    # 创建一个新的工作表
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        count_data.to_excel(writer, sheet_name=output_name, index=False)
    

    print(f"🎉 新的'{output_name}'统计结果已写入！")
